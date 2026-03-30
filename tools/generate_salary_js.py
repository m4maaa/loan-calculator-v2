from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl


LEVEL_ORDER = ["ป.1", "ป.2", "ป.3", "น.1", "น.2", "น.3", "น.4"]


def stringify_step(value):
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def parse_workbook(xlsx_path: Path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    levels = []
    for col in range(1, ws.max_column + 1, 2):
        level_code = ws.cell(1, col).value
        if not level_code:
            continue

        salaries = {}
        for row in range(2, ws.max_row + 1):
            step = ws.cell(row, col).value
            salary = ws.cell(row, col + 1).value
            if step is None or salary is None:
                continue
            salaries[stringify_step(step)] = int(salary)

        levels.append({
            "code": str(level_code).strip(),
            "label": str(level_code).strip(),
            "steps": list(salaries.keys()),
            "salaries": salaries,
        })

    levels.sort(key=lambda item: LEVEL_ORDER.index(item["code"]) if item["code"] in LEVEL_ORDER else 999)
    return levels


def build_salary_js(levels):
    payload = json.dumps(levels, ensure_ascii=False, indent=2)
    return f"""export const SALARY_LEVELS = {payload};

export const SALARY_LEVEL_OPTIONS = SALARY_LEVELS.map((item) => ({{
  value: item.code,
  label: item.label,
}}));

export function getSalaryStepsByLevel(levelCode) {{
  const level = SALARY_LEVELS.find((item) => item.code === levelCode);
  return level?.steps ?? [];
}}

export function getSalaryByLevelAndStep(levelCode, step) {{
  const level = SALARY_LEVELS.find((item) => item.code === levelCode);
  if (!level) return null;
  return level.salaries[String(step)] ?? null;
}}
"""


def main():
    if len(sys.argv) < 2:
        print("Usage: python tools/generate_salary_js.py <xlsx_path> [output_js]")
        raise SystemExit(1)

    xlsx_path = Path(sys.argv[1]).resolve()
    output_js = Path(sys.argv[2]).resolve() if len(sys.argv) >= 3 else Path(__file__).resolve().parents[1] / "data" / "salary.js"

    levels = parse_workbook(xlsx_path)
    output_js.write_text(build_salary_js(levels), encoding="utf-8")
    print(f"Generated {output_js} from {xlsx_path}")


if __name__ == "__main__":
    main()
